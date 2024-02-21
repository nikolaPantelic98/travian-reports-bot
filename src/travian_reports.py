from time import sleep

from openpyxl.reader.excel import load_workbook
from requests import RequestException
from selenium import webdriver
from selenium.common import NoSuchElementException, WebDriverException, NoSuchWindowException
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import random
from datetime import datetime, date, timedelta
import subprocess
import requests
import logging
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import mysql.connector
import re

name = "ReportsLogger"
logger = logging.getLogger(name)
logger.setLevel(logging.INFO)

log_path = os.environ.get('TRAVIAN_REPORTS_BOT_LOG_PATH')
os.makedirs(log_path, exist_ok=True)
log_file = os.path.join(log_path, datetime.now().strftime('%Y-%m-%d') + ".log")
handler = logging.FileHandler(log_file)

logger.addHandler(handler)


def log(message):
    logger.info(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {message}")


def send_telegram_message(message):
    bot_token = os.environ.get('TRAVIAN_REPORTS_BOT_TELEGRAM_MESSAGE_TOKEN')
    chat_id = os.environ.get('TRAVIAN_REPORTS_BOT_TELEGRAM_MESSAGE_CHAT_ID')
    send_text = 'https://api.telegram.org/bot' + bot_token + '/sendMessage?chat_id=' + chat_id + '&parse_mode=Markdown&text=' + message

    response = requests.get(send_text)

    return response.json()


def is_connected():
    try:
        requests.get('https://google.com', timeout=1)
        return True
    except requests.ConnectionError as e:
        log(f"A network error occurred: {e}")
        return False


def restart_wifi():
    subprocess.run(["nmcli", "radio", "wifi", "off"])
    sleep(2)
    subprocess.run(["nmcli", "radio", "wifi", "on"])
    sleep(10)


def setup_driver():
    chrome_options = Options()
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--disable-popup-blocking")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    driver.maximize_window()
    log("Starting script...")
    return driver


def login(driver, username, password):
    driver.get(os.environ.get('TRAVIAN_REPORTS_BOT_REPORTS_URL'))
    sleep(random.uniform(2, 4))

    username_field = driver.find_element(By.NAME, 'name')
    password_field = driver.find_element(By.NAME, 'password')

    for char in username:
        username_field.send_keys(char)
        sleep(random.uniform(0.12, 0.17))

    for char in password:
        password_field.send_keys(char)
        sleep(random.uniform(0.13, 0.21))

    sleep(random.uniform(0.1, 0.3))

    login_button = driver.find_element(By.XPATH, '//button[contains(@class, "textButtonV1")]')
    login_button.click()
    sleep(random.uniform(5.5, 8.5))


def scroll_to(driver, element, scroll_time, scroll_offset):
    driver.execute_script(f"""
        var element = arguments[0];
        var box = element.getBoundingClientRect();
        var body = document.body;
        var docEl = document.documentElement;
        var scrollTop = window.pageYOffset || docEl.scrollTop || body.scrollTop;
        var clientTop = docEl.clientTop || body.clientTop || 0;
        var top  = box.top +  scrollTop - clientTop - {scroll_offset};
        var currenTop = window.pageYOffset || document.documentElement.scrollTop;
        var start = null;
        requestAnimationFrame(function step(timestamp) {{
            if (!start) start = timestamp;
            var progress = timestamp - start;
            if (currenTop < top) {{
                window.scrollTo(0, ((top - currenTop) * progress / {scroll_time}) + currenTop);
            }} else {{
                window.scrollTo(0, currenTop - ((currenTop - top) * progress / {scroll_time}));
            }}
            if (progress < {scroll_time}) {{
                requestAnimationFrame(step);
            }}
        }});
    """, element)


# pause after scroll
def pause(scroll_time, number1, number2):
    sleep(scroll_time / 1000 + random.uniform(number1, number2))


# functions for database
def connect_to_database():
    cnx = mysql.connector.connect(user='root', password='password', host='127.0.0.1', database='travian')
    cursor = cnx.cursor()
    return cnx, cursor


def create_table(cursor):
    create_table_query = """
    CREATE TABLE IF NOT EXISTS reports (
        id INT AUTO_INCREMENT PRIMARY KEY,
        my_village VARCHAR(255),
        farm_village VARCHAR(255),
        attack_date DATETIME,
        farmed_amount INT,
        farmed_capacity INT,
        type_attack VARCHAR(255)
    )
    """
    cursor.execute(create_table_query)


def insert_data_into_table(cursor, data):
    insert_data_query = """
        INSERT INTO reports (my_village, farm_village, attack_date, farmed_amount, farmed_capacity, type_attack)
        VALUES (%s, %s, %s, %s, %s, %s)
        """
    cursor.execute(insert_data_query, data)


def commit_and_close_connection(cnx, cursor):
    cnx.commit()
    cursor.close()
    cnx.close()


# functions to work with Excel
def create_excel_file():
    excel_file_path = os.environ.get('TRAVIAN_REPORTS_BOT_EXCEL_PATH')
    os.makedirs(excel_file_path, exist_ok=True)
    excel_file_name = os.path.join(excel_file_path, datetime.now().strftime('%Y-%m-%d') + ".xlsx")
    return excel_file_name


def open_or_create_workbook(file_name):
    if os.path.exists(file_name):
        wb = load_workbook(filename=file_name)
    else:
        wb = Workbook()
    return wb


def create_worksheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
    return ws


def set_table_headers(ws, headers):
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')


def get_last_id(ws):
    if ws.max_row > 1:
        return ws.cell(row=ws.max_row, column=1).value
    else:
        return 0


def append_row_to_first_ws(ws, row_data, headers, last_id):
    ws.append(row_data)
    for i, data in enumerate(row_data, start=1):
        if i == 1:
            column = ws.column_dimensions[ws.cell(row=last_id + 1, column=i).column_letter]
            column.alignment = Alignment(horizontal='center')


def update_or_add_row_to_second_ws(ws2, farm_village, total, last_id):
    # check if the village exists in the second table
    for row in range(2, ws2.max_row + 1):
        if ws2.cell(row=row, column=2).value == farm_village:
            # if exists, update total_amount_farmed
            current_total = ws2.cell(row=row, column=3).value
            if current_total is None:
                current_total = 0
            ws2.cell(row=row, column=3, value=current_total + total)
            break
    else:
        # if it doesn't exist, add new row
        ws2.cell(row=last_id + 1, column=1, value=last_id)
        ws2.cell(row=last_id + 1, column=1).alignment = Alignment(horizontal='center')
        ws2.cell(row=last_id + 1, column=2, value=farm_village)
        ws2.cell(row=last_id + 1, column=3, value=total)
        last_id += 1
    return last_id


def set_column_width(ws, column_letter, padding):
    max_length = 0
    column = ws[column_letter]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws.column_dimensions[column_letter].width = max_length + padding


def set_column_alignment(ws, column_letter, alignment):
    if alignment not in ['left', 'center', 'right']:
        raise ValueError("Alignment must be 'left', 'center', or 'right'")
    alignment_mapping = {'left': 'left', 'center': 'center', 'right': 'right'}
    cells = ws[column_letter]
    for cell in cells[1:]:
        cell.alignment = Alignment(horizontal=alignment_mapping[alignment])


def read_reports(driver):
    driver.get(os.environ.get('TRAVIAN_REPORTS_BOT_REPORTS_URL'))
    sleep(random.uniform(1.5, 2.5))

    div_tab_attack = driver.find_element(By.XPATH, './/div[contains(@class, "favorKeyoffensive")]')
    a_tab_attack = div_tab_attack.find_element(By.XPATH, './/a')
    a_tab_attack.click()
    sleep(random.uniform(2.1, 3.5))

    # path for Excel file
    excel_file_name = create_excel_file()
    # open or create the Excel document
    wb = open_or_create_workbook(excel_file_name)

    # prepare the first sheet/table
    headers = ["id", "my_village", "farm_village", "attack_date", "farmed_amount", "farmed_capacity", "type_attack"]
    ws = create_worksheet(wb, 'Reports')
    set_table_headers(ws, headers)
    last_id = get_last_id(ws)

    total_farmed = {}

    # database connection to store reports
    cnx, cursor = connect_to_database()
    create_table(cursor)

    # find every report that hasn't been read
    tr_elements = driver.find_elements(By.XPATH, '//tr[td[contains(@class, "newMessage")]]')

    # iterate through each report reversed
    for tr in reversed(tr_elements):
        scroll_time = random.uniform(0.1, 0.12) * 1000
        scroll_offset = random.randint(200, 210)
        # scroll to the next report
        scroll_to(driver, tr, scroll_time, scroll_offset)

        td_new_message = tr.find_element(By.XPATH, './/td[contains(@class, "newMessage")]')
        a_tag_report = td_new_message.find_element(By.XPATH, './/div[@class=""]/a')

        # extract the names of the villages
        villages = re.split(' пљачка | напада ', a_tag_report.text)
        my_village = villages[0]
        farm_village = villages[1]

        # extract the resources
        img_tag_resources = td_new_message.find_element(By.XPATH, './/a[contains(@class, "reportInfoIcon")]/img')
        farmed_amount, farmed_capacity = img_tag_resources.get_attribute('alt').split('/')

        # extract date of the attack
        td_date = tr.find_element(By.XPATH, './/td[contains(@class, "dat")]')
        attack_date = td_date.text
        if 'данас' in attack_date:
            today = date.today()
            attack_date = attack_date.replace('данас', today.strftime("%Y-%m-%d"))
        if 'јуче' in attack_date:
            yesterday = date.today() - timedelta(days=1)
            attack_date = attack_date.replace('јуче', yesterday.strftime("%Y-%m-%d"))

        # extract type of the attack
        img_tag_attack = td_new_message.find_element(By.XPATH, './/img[contains(@class, "iReport")]')
        if img_tag_attack.get_attribute('alt') == 'Победио као нападач без губитака.':
            type_attack = 'green'
        elif img_tag_attack.get_attribute('alt') == 'Победио као нападач са губицима.':
            type_attack = 'orange'
        elif img_tag_attack.get_attribute('alt') == 'Izgubio kao napadac sa gubicima':
            type_attack = 'red'
        else:
            type_attack = 'unknown'

        # find and check the checkbox
        td_checkbox = tr.find_element(By.XPATH, './/td[contains(@class, "sel")]/input')
        pause(scroll_time, 0.01, 0.03)
        td_checkbox.click()

        # fill the Excel data for the first sheet/table
        last_id += 1
        row_data = [last_id, my_village, farm_village, attack_date, farmed_amount, farmed_capacity, type_attack]
        append_row_to_first_ws(ws, row_data, headers, last_id)

        if farm_village in total_farmed:
            total_farmed[farm_village] += int(farmed_amount)
        else:
            total_farmed[farm_village] = int(farmed_amount)

        # insert data to the database table
        data = (my_village, farm_village, attack_date, farmed_amount, farmed_capacity, type_attack)
        insert_data_into_table(cursor, data)

        sleep(random.uniform(0.1, 0.12))

    # find the button to mark the reports as read
    button_wrapper = driver.find_element(By.XPATH, '//div[contains(@class, "buttonWrapper")]')
    scroll_time = random.uniform(0.17, 0.25) * 1000
    scroll_offset = random.randint(180, 190)
    # scroll to the button
    scroll_to(driver, button_wrapper, scroll_time, scroll_offset)
    pause(scroll_time, 0.23, 0.30)
    # click the button
    first_button = button_wrapper.find_element(By.XPATH, './/button')
    first_button.click()

    # set width and alignment for the columns of the first sheet
    set_column_width(ws, 'A', 7)
    set_column_alignment(ws, 'A', 'center')
    set_column_width(ws, 'B', 3)
    set_column_alignment(ws, 'B', 'left')
    set_column_width(ws, 'C', 3)
    set_column_alignment(ws, 'C', 'left')
    set_column_width(ws, 'D', 3)
    set_column_alignment(ws, 'D', 'center')
    set_column_width(ws, 'E', 3)
    set_column_alignment(ws, 'E', 'right')
    set_column_width(ws, 'F', 3)
    set_column_alignment(ws, 'F', 'right')
    set_column_width(ws, 'G', 3)
    set_column_alignment(ws, 'G', 'center')

    # prepare the second sheet/table
    new_headers = ["id", "farm_village", "total_amount_farmed"]
    ws2 = create_worksheet(wb, 'Total_per_villages')
    set_table_headers(ws2, new_headers)

    last_id_value = ws2.cell(row=ws2.max_row, column=1).value if ws2.max_row > 1 else None
    last_id = 1 if not last_id_value else last_id_value + 1

    for farm_village, total in total_farmed.items():
        last_id = update_or_add_row_to_second_ws(ws2, farm_village, total, last_id)

    # set width and alignment for the columns
    set_column_width(ws2, 'A', 7)
    set_column_alignment(ws2, 'A', 'center')
    set_column_width(ws2, 'B', 3)
    set_column_alignment(ws2, 'B', 'left')
    set_column_width(ws2, 'C', 3)
    set_column_alignment(ws2, 'C', 'right')

    # prepare the third sheet/table
    ws3 = create_worksheet(wb, 'Total')
    set_table_headers(ws3, ['total'])

    # calculate the total farmed amount
    total_farmed_amount = sum(total_farmed.values())

    # if the total cell already exists, update it; otherwise, create it
    if ws3['A2'].value is None:
        ws3['A2'] = total_farmed_amount
    else:
        ws3['A2'] = ws3['A2'].value + total_farmed_amount

    # set width and alignment for the columns
    set_column_width(ws3, 'A', 3)
    set_column_alignment(ws3, 'A', 'right')

    # save the Excel file
    wb.save(excel_file_name)

    # commit changes to the database and close the connection
    commit_and_close_connection(cnx, cursor)

    sleep(random.uniform(3.1, 4.1))
    driver.get('https://google.com')
    log("Script executed!")
    send_telegram_message(f"Reports have been successfully read and forwarded ({len(tr_elements)})")
    sleep(random.uniform(680, 940))


# Setup driver
driver = setup_driver()

# Login
while True:
    try:
        login(driver, os.environ.get('TRAVIAN_REPORTS_BOT_USERNAME'), os.environ.get('TRAVIAN_REPORTS_BOT_PASSWORD'))
        break
    except requests.exceptions.ConnectionError as e:
        log(f"A network error occurred when sending a Telegram message: {e}")
        restart_wifi()
        continue
    except RequestException as e:
        log(f"A network error occurred in login form: {e}")
        restart_wifi()
        continue
    except NoSuchElementException as e:
        log(f"A NoSuchElementException occurred in login form: {e}")
        continue
    except Exception as e:
        log(f"An unexpected error form occurred in login: {e}")
        restart_wifi()
        driver.quit()
        driver = setup_driver()
        continue

# Start script
while True:
    try:
        if is_connected():
            read_reports(driver)
        else:
            restart_wifi()
    except RequestException as e:
        log(f"A network error occurred: {e}")
        restart_wifi()
        continue
    except NoSuchElementException as e:
        log(f"A NoSuchElementException occurred: {e}")
        continue
    except NoSuchWindowException as e:
        log(f"A NoSuchWindowException occurred: {e}")
        restart_wifi()
        driver.quit()
        driver = setup_driver()
        login(driver, os.environ.get('TRAVIAN_REPORTS_BOT_USERNAME'), os.environ.get('TRAVIAN_REPORTS_BOT_PASSWORD'))
        continue
    except WebDriverException as e:
        log(f"A WebDriverException occurred: {e}")
        restart_wifi()
        driver.quit()
        driver = setup_driver()
        login(driver, os.environ.get('TRAVIAN_REPORTS_BOT_USERNAME'), os.environ.get('TRAVIAN_REPORTS_BOT_PASSWORD'))
        continue
    except Exception as e:
        log(f"An unexpected error occurred: {e}")
        restart_wifi()
        driver.quit()
        driver = setup_driver()
        login(driver, os.environ.get('TRAVIAN_REPORTS_BOT_USERNAME'), os.environ.get('TRAVIAN_REPORTS_BOT_PASSWORD'))
        continue
